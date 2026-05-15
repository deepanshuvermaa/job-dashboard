"""FastAPI main application with FULL automation integration"""
from fastapi import FastAPI, HTTPException, BackgroundTasks, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
import sys
import os
from datetime import datetime
import shutil
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.db_helper import db
from modules.automation_workflow import AutomationWorkflow
from modules.excel_export_service import ExcelExportService
from modules.github_service import GitHubService
from fastapi.responses import FileResponse

app = FastAPI(title="LinkedIn Automation Suite - FULL VERSION", version="2.0.0")

# Initialize automation workflow
workflow = AutomationWorkflow()

# Initialize Excel exporter
excel_exporter = ExcelExportService()

# Initialize GitHub service
github_service = GitHubService()

# In-memory job cache (stores last search results)
job_cache = []

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== Models ====================

class ContentSourceRequest(BaseModel):
    repo_url: str

class ContentSourceUpdate(BaseModel):
    is_active: Optional[bool] = None

class JobApplicationRequest(BaseModel):
    job_url: str
    job_id: Optional[str] = None

class PostUpdate(BaseModel):
    status: Optional[str] = None
    published_at: Optional[str] = None
    linkedin_url: Optional[str] = None

class SettingsUpdate(BaseModel):
    linkedin_email: Optional[str] = None
    github_username: Optional[str] = None
    openai_api_key: Optional[str] = None
    gemini_api_key: Optional[str] = None
    anthropic_api_key: Optional[str] = None
    auto_post_enabled: Optional[bool] = None
    auto_apply_enabled: Optional[bool] = None
    max_applications_per_day: Optional[int] = None

class AutomationRequest(BaseModel):
    action: str  # 'sync_repos', 'publish_posts', 'apply_jobs', 'full'
    keywords: Optional[str] = "Software Engineer"
    max_applications: Optional[int] = 5

class PortalScanRequest(BaseModel):
    keyword_filter: Optional[str] = None
    location_filter: Optional[str] = None

# ==================== Health ====================

@app.get("/")
async def root():
    return {
        "status": "online",
        "version": "2.0.0 - FULL AUTOMATION",
        "services": ["content_engine", "job_applier", "github_sync", "ai_generation"],
        "message": "LinkedIn Automation Suite API with REAL automation!"
    }

@app.get("/health")
async def health():
    return {"status": "healthy", "backend": "running", "automation": "enabled"}

# ==================== Dashboard ====================

@app.get("/api/dashboard/stats")
async def get_dashboard_stats():
    """Get dashboard statistics"""
    try:
        stats = db.get_dashboard_stats()
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== Settings ====================

@app.get("/api/settings")
async def get_settings():
    """Get all settings"""
    try:
        settings = db.get_all_settings()
        return {
            "linkedin_email": settings.get("linkedin_email", ""),
            "github_username": settings.get("github_username", ""),
            "openai_api_key": settings.get("openai_api_key", ""),
            "gemini_api_key": settings.get("gemini_api_key", ""),
            "anthropic_api_key": settings.get("anthropic_api_key", ""),
            "auto_post_enabled": settings.get("auto_post_enabled", "false") == "true",
            "auto_apply_enabled": settings.get("auto_apply_enabled", "false") == "true",
            "max_applications_per_day": int(settings.get("max_applications_per_day", "50"))
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/settings")
async def update_settings(settings: SettingsUpdate):
    """Update settings"""
    try:
        settings_dict = settings.dict(exclude_unset=True)
        for key, value in settings_dict.items():
            if isinstance(value, bool):
                value = "true" if value else "false"
            db.set_setting(key, str(value))
        return {"success": True, "message": "Settings updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== Content Sources ====================

@app.get("/api/content/sources")
async def get_content_sources():
    """Get all content sources"""
    try:
        sources = db.get_content_sources()
        return {"sources": sources}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/content/sources")
async def add_content_source(request: ContentSourceRequest):
    """Add new content source"""
    try:
        repo_url = request.repo_url.strip()
        if "github.com" in repo_url:
            parts = repo_url.rstrip('/').split('/')
            repo_name = f"{parts[-2]}/{parts[-1]}"
        else:
            repo_name = repo_url
            repo_url = f"https://github.com/{repo_url}"
        source_id = db.add_content_source(repo_name, repo_url)
        return {"success": True, "source_id": source_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/content/sources/{source_id}")
async def update_content_source(source_id: int, update: ContentSourceUpdate):
    """Update content source"""
    try:
        update_data = update.dict(exclude_unset=True)
        db.update_content_source(source_id, **update_data)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/content/sources/{source_id}")
async def delete_content_source(source_id: int):
    """Delete content source"""
    try:
        db.delete_content_source(source_id)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/content/sources/{source_id}/sync")
async def sync_content_source(source_id: int, background_tasks: BackgroundTasks):
    """Sync content from source - REAL GITHUB INTEGRATION"""
    try:
        # Get source info
        sources = db.get_content_sources()
        source = next((s for s in sources if s['id'] == source_id), None)

        if not source:
            raise HTTPException(status_code=404, detail="Source not found")

        # Run sync in background
        background_tasks.add_task(
            workflow.sync_repository_and_generate_posts,
            source_id,
            source['repo_name']
        )

        return {
            "success": True,
            "message": f"Syncing {source['repo_name']} in background..."
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/content/sources/{source_id}/commits")
async def get_repo_commits(source_id: int, limit: int = 30):
    """Get last N commits for a repository (default: 30 for latest commits)"""
    try:
        # Get source info
        sources = db.get_content_sources()
        source = next((s for s in sources if s['id'] == source_id), None)

        if not source:
            raise HTTPException(status_code=404, detail="Source not found")

        # Get commits from GitHub (get latest commits)
        commits = github_service.get_last_n_commits(source['repo_name'], n=limit)

        return {
            "repo_name": source['repo_name'],
            "commits": commits,
            "count": len(commits)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== Content Posts ====================

@app.get("/api/posts")
async def get_posts(status: str = None):
    """Get posts with optional status filter"""
    try:
        posts = db.get_posts(status=status if status else None)
        return {"posts": posts}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/posts")
async def create_post(post_data: dict):
    """Create a new post"""
    try:
        post_id = db.add_post(
            content=post_data.get('content', ''),
            hooks=post_data.get('hooks', []),
            hashtags=post_data.get('hashtags', []),
            pillar=post_data.get('pillar', 'linkedin_questions'),
            source_id=post_data.get('source_id')
        )

        return {
            "success": True,
            "post_id": post_id,
            "message": "Post created successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/content/queue")
async def get_content_queue():
    """Get pending posts"""
    try:
        posts = db.get_posts(status='pending')
        return {"posts": posts}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/content/published")
async def get_published_posts():
    """Get published posts"""
    try:
        posts = db.get_posts(status='published')
        return {"posts": posts}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/content/posts/{post_id}")
async def update_post(post_id: int, update: PostUpdate):
    """Update post"""
    try:
        update_data = update.dict(exclude_unset=True)
        db.update_post(post_id, **update_data)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== Job Search - REAL LINKEDIN SCRAPING ====================

@app.get("/api/jobs/search")
async def search_jobs(
    keywords: Optional[str] = None,
    location: Optional[str] = None,
    job_type: Optional[str] = None,
    experience_level: Optional[str] = None,
    easy_apply: Optional[bool] = None,
    remote: Optional[bool] = None,
    posted_within: Optional[str] = None,
    max_results: Optional[int] = 100
):
    """Search for jobs - REAL LINKEDIN SCRAPING with pagination"""
    global job_cache
    try:
        if not keywords:
            return {"jobs": [], "message": "Please provide search keywords"}

        # Use real LinkedIn scraper
        jobs = workflow.job_scraper.search_jobs(
            keywords=keywords,
            location=location or "United States",
            job_type=job_type,
            experience_level=experience_level,
            easy_apply=easy_apply or False,
            remote=remote or False,
            max_results=max_results,
            posted_within=posted_within
        )

        # Tag all LinkedIn jobs with source
        for job in jobs:
            if 'source' not in job:
                job['source'] = 'linkedin'
            if 'scraped_at' not in job:
                job['scraped_at'] = datetime.now().isoformat()

        # Cache jobs for export
        job_cache = jobs

        # SAVE ALL SCRAPED JOBS TO DATABASE IMMEDIATELY
        saved_count = 0
        for job in jobs:
            try:
                # Extract recruiter info if available
                recruiter_info = job.get('recruiter_info', {})

                recruiter_note = ""
                if recruiter_info and recruiter_info.get('name'):
                    recruiter_note = f"Recruiter: {recruiter_info.get('name', '')} | {recruiter_info.get('profile_url', '')}"

                db.add_application(
                    job_title=job.get('title', 'Unknown'),
                    company=job.get('company', 'Unknown'),
                    location=job.get('location', 'Unknown'),
                    job_url=job.get('job_url', ''),
                    status='found',
                    salary_range=job.get('salary'),
                    date_listed=job.get('posted_date'),
                    notes=f"Source: linkedin | {recruiter_note}".strip()
                )
                saved_count += 1
            except Exception as e:
                print(f"[WARN] Could not save job to database: {e}")
                continue

        print(f"[OK] Saved {saved_count}/{len(jobs)} jobs to database")

        # Also save to scan_history for unified feed
        try:
            from modules.deduplication import JobDeduplicator
            deduplicator = JobDeduplicator()
            deduplicator.filter_new_jobs(jobs)  # This marks them as seen in scan_history
        except Exception as e:
            print(f"[WARN] Could not save to scan_history: {e}")

        # DON'T close browser - keep session for Quick Apply
        # workflow.job_scraper.close()

        return {"jobs": jobs, "count": len(jobs), "saved_to_db": saved_count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/jobs/export")
async def export_jobs_to_excel():
    """Export cached job search results to Excel"""
    global job_cache
    try:
        if not job_cache:
            raise HTTPException(status_code=400, detail="No jobs to export. Please search for jobs first.")

        # Generate Excel file
        filepath = excel_exporter.export_jobs_to_excel(job_cache)

        # Return file for download
        return FileResponse(
            path=filepath,
            filename=os.path.basename(filepath),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== Job Applications ====================

@app.post("/api/jobs/apply")
async def apply_to_job(request: JobApplicationRequest, background_tasks: BackgroundTasks):
    """Apply to job - REAL EASY APPLY"""
    try:
        # Apply in background
        def apply_task():
            success = workflow.job_scraper.apply_to_job(request.job_url)
            if success:
                db.add_application(
                    job_title="Applied Job",
                    company="Company",
                    location="Location",
                    job_url=request.job_url
                )

        background_tasks.add_task(apply_task)

        return {"success": True, "message": "Application in progress..."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/jobs/applications")
async def get_applications(status: Optional[str] = None):
    """Get applications"""
    try:
        applications = db.get_applications(status=status)
        return {"applications": applications}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== AUTOMATION ENDPOINTS ====================

@app.post("/api/automation/run")
async def run_automation(request: AutomationRequest, background_tasks: BackgroundTasks):
    """Run automation workflow"""
    try:
        action = request.action

        if action == 'sync_repos':
            background_tasks.add_task(workflow.auto_sync_all_repos)
            message = "Syncing all repositories..."

        elif action == 'publish_posts':
            background_tasks.add_task(workflow.auto_publish_approved_posts)
            message = "Publishing approved posts..."

        elif action == 'apply_jobs':
            background_tasks.add_task(
                workflow.auto_job_search_and_apply,
                request.keywords,
                "United States",
                request.max_applications
            )
            message = f"Applying to {request.max_applications} jobs..."

        elif action == 'full':
            background_tasks.add_task(workflow.run_full_automation)
            message = "Running full automation workflow..."

        else:
            raise HTTPException(status_code=400, detail="Invalid action")

        return {
            "success": True,
            "message": message,
            "action": action
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main_v2:app", host="0.0.0.0", port=8000, reload=True)

# ==================== AI Auto-Messaging ====================

from modules.ai_message_generator import AIMessageGenerator
from modules.linkedin_auto_messenger import LinkedInAutoMessenger

# Initialize services
ai_message_gen = AIMessageGenerator()
auto_messenger_instance = None

class MessageRequest(BaseModel):
    max_messages: Optional[int] = 10
    delay_seconds: Optional[int] = 30

@app.post("/api/messages/generate")
async def generate_messages():
    """Generate personalized AI messages for all cached jobs"""
    global job_cache, ai_message_gen
    try:
        if not job_cache:
            raise HTTPException(status_code=400, detail="No jobs cached. Please search for jobs first.")

        messages = ai_message_gen.generate_bulk_messages(job_cache, max_messages=25)

        return {
            "success": True,
            "count": len(messages),
            "messages": messages
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/messages/send")
async def send_auto_messages(request: MessageRequest, background_tasks: BackgroundTasks):
    """Send AI-generated messages to recruiters"""
    global job_cache, ai_message_gen, auto_messenger_instance
    try:
        if not job_cache:
            raise HTTPException(status_code=400, detail="No jobs cached. Search for jobs first.")

        # Generate messages
        messages = ai_message_gen.generate_bulk_messages(job_cache, max_messages=request.max_messages)

        if not messages:
            raise HTTPException(status_code=400, detail="No messages could be generated. Ensure jobs have recruiter info.")

        # Initialize messenger if needed
        if not auto_messenger_instance:
            auto_messenger_instance = LinkedInAutoMessenger()

        # Send messages in background
        def send_messages_task():
            results = auto_messenger_instance.send_bulk_messages(
                messages,
                delay_seconds=request.delay_seconds,
                max_messages=request.max_messages
            )
            print(f"Auto-messaging complete: {results['sent']} sent, {results['failed']} failed")

        background_tasks.add_task(send_messages_task)

        return {
            "success": True,
            "message": f"Sending {len(messages)} personalized messages in background...",
            "preview": messages[:3]  # Show first 3 as preview
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== GitHub All Repos ====================

@app.get("/api/github/all-repos")
async def get_all_user_repos(username: Optional[str] = None):
    """Get all GitHub repositories for the authenticated user"""
    try:
        # Get username from settings if not provided
        if not username:
            settings = db.get_all_settings()
            username = settings.get("github_username", "")
            if not username:
                raise HTTPException(status_code=400, detail="GitHub username not configured in settings")

        # Fetch all repos
        repos = github_service.get_user_repos(username)

        return {
            "success": True,
            "username": username,
            "repos": repos,
            "count": len(repos)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/github/repos/add-multiple")
async def add_multiple_repos(repo_names: List[str]):
    """Add multiple repositories at once"""
    try:
        added = []
        for repo_name in repo_names:
            repo_url = f"https://github.com/{repo_name}"
            try:
                source_id = db.add_content_source(repo_name, repo_url)
                added.append({"repo_name": repo_name, "source_id": source_id})
            except Exception as e:
                print(f"Error adding {repo_name}: {e}")
                continue

        return {
            "success": True,
            "added_count": len(added),
            "repos": added
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== Create Post from Keywords ====================

from modules.ai_content_generator import AIContentGenerator
from modules.linkedin_post_generator import LinkedInPostGenerator

ai_content_gen = AIContentGenerator()
linkedin_post_gen = LinkedInPostGenerator()

class KeywordPostRequest(BaseModel):
    keywords: str
    tone: Optional[str] = "professional"  # professional, casual, technical, inspirational
    length: Optional[str] = "medium"  # short, medium, long
    include_hashtags: Optional[bool] = True
    include_emoji: Optional[bool] = False

@app.post("/api/content/create-from-keywords")
async def create_post_from_keywords(request: KeywordPostRequest):
    """Generate LinkedIn post from keywords using AI"""
    try:
        # Generate post using AI
        post_content = ai_content_gen.generate_post_from_keywords(
            keywords=request.keywords,
            tone=request.tone,
            length=request.length,
            include_hashtags=request.include_hashtags,
            include_emoji=request.include_emoji
        )

        # Save to database as draft
        post_id = db.add_post(
            content=post_content,
            hooks=[],
            hashtags=[],
            pillar="keywords",
            source_id=None
        )

        return {
            "success": True,
            "post_id": post_id,
            "content": post_content,
            "message": "Post created successfully! Review it in the Content Queue."
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== GitHub Commit-Based Post Generation ====================

@app.post("/api/content/generate-from-commits")
async def generate_posts_from_commits(background_tasks: BackgroundTasks):
    """Generate posts from all active GitHub repositories"""
    try:
        # Get all active sources
        sources = db.get_content_sources()
        active_sources = [s for s in sources if s.get('is_active', True)]

        if not active_sources:
            raise HTTPException(status_code=400, detail="No active repositories found. Add repositories first.")

        # Generate posts in background
        def generate_posts_task():
            total_generated = 0
            for source in active_sources:
                try:
                    # Get recent commits
                    activity = github_service.get_repo_activity_summary(source['repo_name'], days=7)

                    if activity['commit_count'] == 0:
                        continue

                    # Generate post from commits
                    post_content = ai_content_gen.generate_post_from_commits(
                        repo_name=source['repo_name'],
                        commits=activity['commits']
                    )

                    if post_content:
                        # Save to database
                        db.add_post(
                            content=post_content,
                            hooks=[],
                            hashtags=[],
                            pillar="github",
                            source_id=source['id']
                        )
                        total_generated += 1

                        # Update last_synced
                        db.update_content_source(source['id'], last_synced=datetime.now().isoformat())
                except Exception as e:
                    print(f"Error generating post for {source['repo_name']}: {e}")
                    continue

            print(f"Generated {total_generated} posts from GitHub commits")

        background_tasks.add_task(generate_posts_task)

        return {
            "success": True,
            "message": f"Generating posts from {len(active_sources)} repositories...",
            "repo_count": len(active_sources)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/content/sources/{source_id}/generate-post")
async def generate_post_for_repo(source_id: int):
    """Generate a LinkedIn post for a specific repository"""
    try:
        # Get source info
        sources = db.get_content_sources()
        source = next((s for s in sources if s['id'] == source_id), None)

        if not source:
            raise HTTPException(status_code=404, detail="Repository not found")

        # Get recent commits (check all commits, not just last 7 days)
        commits = github_service.get_last_n_commits(source['repo_name'], n=5)

        if not commits or len(commits) == 0:
            raise HTTPException(status_code=400, detail="No commits found in this repository")

        # Generate post from commits + README
        post_content = ai_content_gen.generate_post_from_commits(
            repo_name=source['repo_name'],
            commits=commits,
            include_readme=True
        )

        if not post_content:
            raise HTTPException(status_code=500, detail="Failed to generate post content")

        # Save to database
        post_id = db.add_post(
            content=post_content,
            hooks=[],
            hashtags=[],
            pillar="github",
            source_id=source['id']
        )

        # Update last_synced
        db.update_content_source(source_id, last_synced=datetime.now().isoformat())

        return {
            "success": True,
            "post_id": post_id,
            "content": post_content,
            "message": "Post generated successfully! Check the Content Queue to review it."
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error generating post for repo {source_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== LinkedIn Post Generator (50+ Posts) ====================

class BulkPostRequest(BaseModel):
    num_posts: Optional[int] = 50
    industry: Optional[str] = "software development"
    achievements: Optional[str] = ""
    notes: Optional[str] = ""
    tone: Optional[str] = "conversational"
    length: Optional[str] = "medium"
    selected_questions: Optional[List[int]] = None

@app.get("/api/content/questions")
async def get_linkedin_questions():
    """Get all 30 LinkedIn question prompts"""
    from modules.linkedin_post_generator import LINKEDIN_QUESTIONS
    return {
        "questions": [{"id": i, "question": q} for i, q in enumerate(LINKEDIN_QUESTIONS)],
        "total": len(LINKEDIN_QUESTIONS)
    }

@app.post("/api/content/generate-bulk-posts")
async def generate_bulk_linkedin_posts(request: BulkPostRequest, background_tasks: BackgroundTasks):
    """Generate 50+ authentic LinkedIn posts from question prompts"""
    try:
        user_context = {
            'industry': request.industry,
            'achievements': request.achievements,
            'notes': request.notes
        }

        # Generate in background
        def generate_posts_task():
            posts = linkedin_post_gen.generate_bulk_posts(
                num_posts=request.num_posts,
                user_context=user_context,
                tone=request.tone,
                length=request.length,
                selected_questions=request.selected_questions
            )

            # Save all posts to database
            for post in posts:
                db.add_post(
                    content=post['content'],
                    hooks=[],
                    hashtags=[],
                    pillar="linkedin_questions",
                    source_id=None
                )

            print(f"✓ Saved {len(posts)} posts to database")

        background_tasks.add_task(generate_posts_task)

        return {
            "success": True,
            "message": f"Generating {request.num_posts} LinkedIn posts in background...",
            "estimated_time": f"{request.num_posts * 2} seconds",
            "check_queue": "Posts will appear in Content Queue when ready"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/content/generate-single-post")
async def generate_single_linkedin_post(
    question_id: int,
    industry: Optional[str] = "software development",
    achievements: Optional[str] = "",
    notes: Optional[str] = "",
    tone: Optional[str] = "conversational",
    length: Optional[str] = "medium"
):
    """Generate a single LinkedIn post from a specific question"""
    try:
        from modules.linkedin_post_generator import LINKEDIN_QUESTIONS

        if question_id < 0 or question_id >= len(LINKEDIN_QUESTIONS):
            raise HTTPException(status_code=400, detail="Invalid question ID")

        question = LINKEDIN_QUESTIONS[question_id]

        user_context = {
            'industry': industry,
            'achievements': achievements,
            'notes': notes
        }

        post_content = linkedin_post_gen.generate_post_from_question(
            question=question,
            user_context=user_context,
            tone=tone,
            length=length
        )

        # Save to database
        post_id = db.add_post(
            content=post_content,
            hooks=[],
            hashtags=[],
            pillar="linkedin_questions",
            source_id=None
        )

        return {
            "success": True,
            "post_id": post_id,
            "question": question,
            "content": post_content,
            "word_count": len(post_content.split()),
            "char_count": len(post_content)
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ===========================
# LINKEDIN AUTO-POSTING
# ===========================

from modules.linkedin_auto_poster import LinkedInAutoPoster

# Global auto-poster instance
auto_poster = None


@app.post("/api/linkedin/auto-post/start")
async def start_auto_posting():
    """Initialize LinkedIn auto-poster and login"""
    global auto_poster

    try:
        auto_poster = LinkedInAutoPoster()
        auto_poster.initialize_driver()

        if auto_poster.login_to_linkedin():
            return {
                "success": True,
                "message": "Successfully logged into LinkedIn. Ready to post."
            }
        else:
            return {
                "success": False,
                "message": "Failed to login to LinkedIn. Please check credentials."
            }

    except Exception as e:
        return {
            "success": False,
            "message": f"Error initializing auto-poster: {str(e)}"
        }


@app.post("/api/linkedin/auto-post/post-single")
async def auto_post_single(request: dict):
    """Post a single piece of content to LinkedIn"""
    global auto_poster

    try:
        if not auto_poster:
            return {
                "success": False,
                "message": "Auto-poster not initialized. Call /api/linkedin/auto-post/start first."
            }

        post_id = request.get('post_id')
        content = request.get('content')

        if not content:
            return {
                "success": False,
                "message": "No content provided"
            }

        result = auto_poster.create_post(content)

        # Update post status in database if post_id provided
        if result['success'] and post_id:
            db.update_post_status(post_id, 'published')

        return result

    except Exception as e:
        return {
            "success": False,
            "message": f"Error posting: {str(e)}"
        }


@app.post("/api/linkedin/auto-post/post-multiple")
async def auto_post_multiple(request: dict):
    """Post multiple pieces of content to LinkedIn"""
    global auto_poster

    try:
        if not auto_poster:
            return {
                "success": False,
                "message": "Auto-poster not initialized. Call /api/linkedin/auto-post/start first."
            }

        post_ids = request.get('post_ids', [])
        delay_between = request.get('delay_between', 10)

        # Get posts from database
        posts_content = []
        for post_id in post_ids:
            post = db.get_post_by_id(post_id)
            if post:
                posts_content.append(post['content'])

        if not posts_content:
            return {
                "success": False,
                "message": "No posts found with provided IDs"
            }

        result = auto_poster.post_multiple(posts_content, delay_between)

        # Update published posts in database
        if result['successful'] > 0:
            for i, post_id in enumerate(post_ids):
                if i < len(result['results']) and result['results'][i]['success']:
                    db.update_post_status(post_id, 'published')

        return result

    except Exception as e:
        return {
            "success": False,
            "message": f"Error posting multiple: {str(e)}"
        }


@app.post("/api/linkedin/auto-post/stop")
async def stop_auto_posting():
    """Close the auto-poster browser"""
    global auto_poster

    try:
        if auto_poster:
            auto_poster.close()
            auto_poster = None

        return {
            "success": True,
            "message": "Auto-poster stopped and browser closed"
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"Error stopping auto-poster: {str(e)}"
        }


@app.get("/api/linkedin/auto-post/status")
async def get_auto_post_status():
    """Check if auto-poster is active"""
    global auto_poster

    return {
        "active": auto_poster is not None,
        "message": "Auto-poster is ready" if auto_poster else "Auto-poster not initialized"
    }


# ===========================
# SMART RESUME PARSER & PROFILE
# ===========================

from modules.smart_resume_parser import SmartResumeParser

# Global resume parser instance
resume_parser = None
user_profile = {}


@app.post("/api/profile/upload-resume")
async def upload_resume(file: UploadFile = File(...)):
    """Upload and parse resume using AI"""
    global resume_parser, user_profile

    try:
        # Create uploads directory if it doesn't exist
        upload_dir = Path(__file__).parent.parent / "uploads"
        upload_dir.mkdir(exist_ok=True)

        # Save the uploaded file
        file_path = upload_dir / file.filename
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Parse resume using AI
        resume_parser = SmartResumeParser()
        result = resume_parser.parse_resume(str(file_path))

        if result.get('success'):
            user_profile = result['data']

            # Persist to database so it survives redeploys
            try:
                from core.database import SessionLocal
                from models.user import User, UserProfile
                db_session = SessionLocal()
                # Find scraper user or first user
                user = db_session.query(User).first()
                if user:
                    profile = db_session.query(UserProfile).filter(UserProfile.user_id == user.id).first()
                    if not profile:
                        profile = UserProfile(user_id=user.id)
                        db_session.add(profile)
                    profile.resume_data = user_profile
                    profile.headline = user_profile.get("name", "")
                    profile.summary = user_profile.get("summary", "")
                    profile.skills = user_profile.get("skills", [])
                    profile.experience = user_profile.get("experience", [])
                    profile.education = user_profile.get("education", [])
                    db_session.commit()
                    print(f"[Resume] Profile saved to DB for user {user.email}")
                else:
                    print("[Resume] WARNING: No user found in DB to save profile")
                db_session.close()
            except Exception as db_err:
                print(f"[Resume] FAILED to persist profile to DB: {db_err}")
                import traceback
                traceback.print_exc()

            return {
                "success": True,
                "message": "Resume parsed successfully using AI",
                "profile": user_profile
            }
        else:
            return {
                "success": False,
                "message": result.get('error', 'Failed to parse resume')
            }

    except Exception as e:
        return {
            "success": False,
            "message": f"Error uploading resume: {str(e)}"
        }


@app.get("/api/profile/get")
async def get_user_profile():
    """Get the current user profile"""
    global user_profile

    if not user_profile:
        # Load from database
        try:
            from core.database import SessionLocal
            from models.user import User, UserProfile
            db_session = SessionLocal()
            user = db_session.query(User).first()
            if user:
                profile = db_session.query(UserProfile).filter(UserProfile.user_id == user.id).first()
                if profile and profile.resume_data:
                    user_profile = profile.resume_data
            db_session.close()
        except Exception as e:
            print(f"[WARN] Failed to load profile from DB: {e}")

    if not user_profile:
        return {
            "success": False,
            "message": "No profile data available. Please upload your resume first."
        }

    return {
        "success": True,
        "profile": user_profile
    }


@app.post("/api/profile/update")
async def update_user_profile(updates: dict):
    """Update user profile with manual edits"""
    global user_profile, resume_parser

    try:
        if not user_profile:
            return {
                "success": False,
                "message": "No profile exists. Upload resume first."
            }

        # Update the profile
        for key, value in updates.items():
            if key in user_profile:
                user_profile[key] = value
            else:
                user_profile[key] = value

        # Update resume parser if it exists
        if resume_parser:
            resume_parser.structured_data = user_profile

        return {
            "success": True,
            "message": "Profile updated successfully",
            "profile": user_profile
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"Error updating profile: {str(e)}"
        }


@app.get("/api/profile/for-content-generation")
async def get_profile_for_content():
    """Get formatted profile data for LinkedIn post generation"""
    global resume_parser, user_profile

    try:
        if not user_profile:
            return {
                "success": False,
                "message": "No profile data available"
            }

        if resume_parser:
            formatted_profile = resume_parser.get_profile_for_content_generation()
        else:
            # Create basic formatted profile from user_profile
            formatted_profile = {
                'name': user_profile.get('name', ''),
                'industry': user_profile.get('primary_industry', 'Technology'),
                'experience_years': user_profile.get('total_experience_years', 1),
                'summary': user_profile.get('summary', ''),
                'top_skills': user_profile.get('skills', {}).get('programming', [])[:5],
                'achievements': user_profile.get('achievements', [])[:5],
                'full_data': user_profile
            }

        return {
            "success": True,
            "profile": formatted_profile
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"Error formatting profile: {str(e)}"
        }


# ==================== Easy Apply Automation ====================

from config.user_config import UserConfig
from modules.easy_apply_bot import EasyApplyBot
from modules.ai_providers import AIProviderManager

# Global instances
user_config_instance = None
easy_apply_bot = None


class UserConfigUpdate(BaseModel):
    """Request model for user configuration updates"""
    section: str  # personal, professional, education, work_status, compensation, application, linkedin, ai, job_search
    updates: dict


class EasyApplyJobsRequest(BaseModel):
    """Request model for Easy Apply automation"""
    keywords: str = "Software Engineer"
    location: str = "United States"
    max_applications: int = 10
    easy_apply_only: bool = True
    use_advanced_bot: bool = True


@app.get("/api/config/user")
async def get_user_config():
    """Get current user configuration"""
    global user_config_instance

    try:
        if not user_config_instance:
            user_config_instance = UserConfig()

        return {
            "success": True,
            "config": user_config_instance.get_all()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/config/user")
async def update_user_config(request: UserConfigUpdate):
    """Update user configuration"""
    global user_config_instance

    try:
        if not user_config_instance:
            user_config_instance = UserConfig()

        # Update the specified section
        user_config_instance.update_section(request.section, request.updates)
        user_config_instance.save()

        return {
            "success": True,
            "message": f"Updated {request.section} configuration",
            "config": user_config_instance.get_all()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/validate")
async def validate_user_config():
    """Validate user configuration"""
    global user_config_instance

    try:
        if not user_config_instance:
            user_config_instance = UserConfig()

        issues = user_config_instance.validate()

        return {
            "success": True,
            "is_valid": user_config_instance.is_valid(),
            "issues": issues
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/jobs/easy-apply/start")
async def start_easy_apply_automation(request: EasyApplyJobsRequest, background_tasks: BackgroundTasks):
    """Start Easy Apply automation with intelligent question answering"""
    global user_config_instance, easy_apply_bot

    try:
        # Load user config
        if not user_config_instance:
            user_config_instance = UserConfig()

        # Validate config
        if not user_config_instance.is_valid():
            issues = user_config_instance.validate()
            return {
                "success": False,
                "error": "User configuration is incomplete",
                "issues": issues
            }

        # Get flat config for bot
        flat_config = user_config_instance.get_flat_config()

        # Initialize AI provider
        ai_config = user_config_instance.get_section('ai')
        ai_manager = AIProviderManager(ai_config)
        ai_client = ai_manager.get_raw_client()

        # Initialize Easy Apply bot
        easy_apply_bot = EasyApplyBot(flat_config, ai_client)
        easy_apply_bot.initialize_driver()

        # Login to LinkedIn
        linkedin_email = user_config_instance.get('linkedin.email')
        linkedin_password = user_config_instance.get('linkedin.password')

        if not easy_apply_bot.login_to_linkedin(linkedin_email, linkedin_password):
            return {
                "success": False,
                "error": "Failed to login to LinkedIn"
            }

        # Search for jobs (using existing scraper)
        from modules.linkedin_job_scraper import LinkedInJobScraper

        scraper = LinkedInJobScraper(linkedin_email, linkedin_password)
        scraper.driver = easy_apply_bot.driver  # Reuse driver
        scraper.logged_in = True

        jobs = scraper.search_jobs(
            keywords=request.keywords,
            location=request.location,
            easy_apply=request.easy_apply_only,
            max_results=request.max_applications
        )

        if not jobs:
            return {
                "success": False,
                "error": "No jobs found matching criteria"
            }

        # Filter to Easy Apply jobs only
        easy_apply_jobs = [job for job in jobs if job.get('easy_apply', False)]

        if not easy_apply_jobs:
            return {
                "success": False,
                "error": "No Easy Apply jobs found"
            }

        # Apply to jobs in background
        def apply_to_jobs_task():
            job_urls = [job['job_url'] for job in easy_apply_jobs]
            results = easy_apply_bot.apply_to_multiple_jobs(job_urls, delay_between=5)

            # Save results to database
            for i, result in enumerate(results):
                if i < len(easy_apply_jobs):
                    job = easy_apply_jobs[i]

                    db.add_application(
                        job_title=job.get('title', 'Unknown'),
                        company=job.get('company', 'Unknown'),
                        location=job.get('location', 'Unknown'),
                        job_url=result['job_url'],
                        application_success=result['success'],
                        error_message=result.get('error'),
                        screenshot_path=result.get('screenshot_path'),
                        questions_asked=result.get('questions_asked', []),
                        status='applied' if result['success'] else 'failed'
                    )

            print(f"✓ Applied to {len(results)} jobs")
            easy_apply_bot.close()

        background_tasks.add_task(apply_to_jobs_task)

        return {
            "success": True,
            "message": f"Applying to {len(easy_apply_jobs)} Easy Apply jobs in background...",
            "jobs_found": len(jobs),
            "easy_apply_jobs": len(easy_apply_jobs)
        }

    except Exception as e:
        if easy_apply_bot:
            easy_apply_bot.close()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/jobs/easy-apply/single")
async def apply_single_job_easy_apply(request: JobApplicationRequest):
    """Apply to a single job using Easy Apply with intelligent question answering"""
    global user_config_instance

    try:
        # Load user config
        if not user_config_instance:
            user_config_instance = UserConfig()

        # Validate config
        if not user_config_instance.is_valid():
            issues = user_config_instance.validate()
            return {
                "success": False,
                "error": "User configuration is incomplete",
                "issues": issues
            }

        # Get flat config
        flat_config = user_config_instance.get_flat_config()

        # Initialize AI provider
        ai_config = user_config_instance.get_section('ai')
        ai_manager = AIProviderManager(ai_config)
        ai_client = ai_manager.get_raw_client()

        # Initialize Easy Apply bot
        bot = EasyApplyBot(flat_config, ai_client)
        bot.initialize_driver()

        # Login
        linkedin_email = user_config_instance.get('linkedin.email')
        linkedin_password = user_config_instance.get('linkedin.password')

        if not bot.login_to_linkedin(linkedin_email, linkedin_password):
            bot.close()
            return {
                "success": False,
                "error": "Failed to login to LinkedIn"
            }

        # Apply to job
        result = bot.apply_to_job(request.job_url)

        # Save to database
        if result['success']:
            db.add_application(
                job_title="Applied Job",
                company="Unknown",
                location="Unknown",
                job_url=request.job_url,
                application_success=True,
                questions_asked=result.get('questions_asked', []),
                screenshot_path=result.get('screenshot_path'),
                status='applied'
            )

        bot.close()

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/jobs/easy-apply/stats")
async def get_easy_apply_stats():
    """Get Easy Apply automation statistics"""
    global easy_apply_bot

    try:
        if not easy_apply_bot:
            return {
                "success": False,
                "message": "No active Easy Apply session"
            }

        stats = easy_apply_bot.get_stats()

        return {
            "success": True,
            "stats": stats
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Multi-Query Search ====================

class MultiSearchRequest(BaseModel):
    queries: List[str]
    location: Optional[str] = "United States"
    max_results_per_query: Optional[int] = 50
    job_type: Optional[str] = None
    experience_level: Optional[str] = None
    easy_apply: Optional[bool] = False
    remote: Optional[bool] = False
    posted_within: Optional[str] = None

@app.post("/api/jobs/search-multi")
async def search_jobs_multi(request: MultiSearchRequest, background_tasks: BackgroundTasks):
    """Search jobs with multiple queries and deduplicate results"""
    global job_cache
    try:
        if not request.queries:
            return {"jobs": [], "message": "Please provide search queries"}

        jobs = workflow.job_scraper.search_jobs_multi(
            queries=request.queries,
            location=request.location or "United States",
            max_results_per_query=request.max_results_per_query,
            job_type=request.job_type,
            experience_level=request.experience_level,
            easy_apply=request.easy_apply or False,
            remote=request.remote or False,
            posted_within=request.posted_within
        )

        job_cache = jobs

        # Save to database
        saved_count = 0
        for job in jobs:
            try:
                recruiter_info = job.get('recruiter_info', {})
                db.add_application(
                    job_title=job.get('title', 'Unknown'),
                    company=job.get('company', 'Unknown'),
                    location=job.get('location', 'Unknown'),
                    job_url=job.get('job_url', ''),
                    status='found',
                    salary_range=job.get('salary'),
                    date_listed=job.get('posted_date'),
                    hr_name=recruiter_info.get('name') if recruiter_info else None,
                    hr_profile_url=recruiter_info.get('profile_url') if recruiter_info else None,
                )
                saved_count += 1
            except Exception as e:
                continue

        return {"jobs": jobs, "count": len(jobs), "saved_to_db": saved_count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Portal Scanner ====================

@app.get("/api/portals/config")
async def get_portal_config():
    """Get portal scanner configuration"""
    try:
        from modules.portal_scanner import PortalScanner
        scanner = PortalScanner()
        return {"config": scanner.config, "success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/portals/config")
async def update_portal_config(config: dict):
    """Update portal scanner configuration"""
    try:
        import yaml
        config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config", "portals.yml")
        with open(config_path, 'w') as f:
            yaml.dump(config, f, default_flow_style=False)
        return {"success": True, "message": "Portal config updated"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Global scan status tracker
portal_scan_status = {"running": False, "progress": {}, "results": []}

@app.post("/api/portals/scan")
async def start_portal_scan(background_tasks: BackgroundTasks, request: PortalScanRequest = None):
    """Start scanning all configured portals"""
    global portal_scan_status
    try:
        if portal_scan_status["running"]:
            return {"success": False, "message": "Scan already in progress"}

        def run_scan():
            global portal_scan_status
            portal_scan_status = {"running": True, "progress": {"status": "starting"}, "results": []}
            try:
                from modules.portal_scanner import PortalScanner
                from modules.deduplication import JobDeduplicator
                scanner = PortalScanner()
                deduplicator = JobDeduplicator()

                jobs = scanner.scan_all_portals(
                    progress_callback=lambda p: portal_scan_status.update({"progress": p}),
                    keyword_filter=request.keyword_filter if request else None,
                    location_filter=request.location_filter if request else None
                )
                new_jobs, skipped = deduplicator.filter_new_jobs(jobs)

                # Save new jobs to database
                saved = 0
                for job in new_jobs:
                    try:
                        db.add_application(
                            job_title=job.get('title', 'Unknown'),
                            company=job.get('company', 'Unknown'),
                            location=job.get('location', 'Unknown'),
                            job_url=job.get('job_url', ''),
                            status='found',
                            salary_range=job.get('salary'),
                            notes=f"Source: {job.get('source', 'portal_scan')}"
                        )
                        saved += 1
                    except:
                        continue

                portal_scan_status = {
                    "running": False,
                    "progress": {"status": "completed"},
                    "results": {
                        "total_scraped": len(jobs),
                        "new_jobs": len(new_jobs),
                        "duplicates_skipped": skipped,
                        "saved_to_db": saved
                    }
                }
            except Exception as e:
                portal_scan_status = {"running": False, "progress": {"status": "error", "error": str(e)}, "results": []}

        background_tasks.add_task(run_scan)
        return {"success": True, "message": "Portal scan started"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/portals/scan/status")
async def get_portal_scan_status():
    """Get current portal scan progress"""
    return portal_scan_status


# ==================== Job Evaluation ====================

class EvaluateJobRequest(BaseModel):
    job_url: Optional[str] = None
    job_data: Optional[dict] = None

class EvaluateBatchRequest(BaseModel):
    job_urls: Optional[List[str]] = None
    min_grade: Optional[str] = None

@app.post("/api/jobs/evaluate")
async def evaluate_job(request: EvaluateJobRequest):
    """Evaluate a single job on 10 dimensions"""
    try:
        from modules.job_evaluator import JobEvaluator
        evaluator = JobEvaluator()

        if request.job_data:
            result = evaluator.evaluate_job(request.job_data)
        elif request.job_url:
            # Find job in cache
            job = next((j for j in job_cache if j.get('job_url') == request.job_url), None)
            if not job:
                raise HTTPException(status_code=404, detail="Job not found in cache")
            result = evaluator.evaluate_job(job)
        else:
            raise HTTPException(status_code=400, detail="Provide job_url or job_data")

        return {"success": True, "evaluation": result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/jobs/evaluate-batch")
async def evaluate_batch(request: EvaluateBatchRequest, background_tasks: BackgroundTasks):
    """Evaluate all cached jobs in batch"""
    try:
        from modules.job_evaluator import JobEvaluator
        evaluator = JobEvaluator()

        jobs_to_eval = job_cache
        if request.job_urls:
            jobs_to_eval = [j for j in job_cache if j.get('job_url') in request.job_urls]

        results = evaluator.evaluate_batch(jobs_to_eval)

        if request.min_grade:
            grade_order = {'A': 5, 'B': 4, 'C': 3, 'D': 2, 'F': 1}
            min_val = grade_order.get(request.min_grade, 0)
            results = [r for r in results if grade_order.get(r.get('evaluation', {}).get('grade', 'F'), 0) >= min_val]

        return {"success": True, "evaluated": len(results), "jobs": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/jobs/evaluations")
async def get_evaluations(min_grade: Optional[str] = None):
    """Get all job evaluations from database"""
    try:
        from modules.job_evaluator import JobEvaluator
        evaluator = JobEvaluator()
        evaluations = evaluator.get_all_evaluations(min_grade=min_grade)
        return {"success": True, "evaluations": evaluations, "count": len(evaluations)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Resume Generator ====================

class GenerateResumeRequest(BaseModel):
    job_url: Optional[str] = None
    job_data: Optional[dict] = None
    job_description: Optional[str] = None  # Raw JD text from textarea
    job_title: Optional[str] = None
    company: Optional[str] = None
    archetype: Optional[str] = "general"
    format: Optional[str] = "Letter"
    inject_keywords: Optional[List[str]] = None

@app.post("/api/resume/generate")
async def generate_resume(request: GenerateResumeRequest):
    """Generate a tailored ATS-optimized resume for a specific job"""
    try:
        from modules.resume_generator.resume_builder import ResumeBuilder
        from modules.resume_generator.keyword_extractor import KeywordExtractor

        builder = ResumeBuilder()
        extractor = KeywordExtractor()

        job = request.job_data

        # Try job_cache first
        if not job and request.job_url:
            job = next((j for j in job_cache if j.get('job_url') == request.job_url), None)

        # Try scan_history database
        if not job and request.job_url:
            try:
                conn = db.get_connection()
                cursor = conn.cursor()
                cursor.execute('SELECT * FROM scan_history WHERE job_url = ?', (request.job_url,))
                row = cursor.fetchone()
                conn.close()
                if row:
                    job = {
                        'title': row['title'],
                        'company': row['company'],
                        'job_url': row['job_url'],
                        'source': row['source'],
                        'description_snippet': '',
                    }
            except:
                pass

        # Build job from raw JD text (when user pastes description manually)
        if not job and request.job_description:
            job = {
                'title': request.job_title or 'Software Developer',
                'company': request.company or 'Company',
                'job_url': request.job_url or '',
                'description_snippet': request.job_description,
                'source': 'manual',
            }

        # Final fallback: if we have a URL but no description, still generate with title
        if not job and (request.job_title or request.company):
            job = {
                'title': request.job_title or 'Position',
                'company': request.company or 'Company',
                'job_url': request.job_url or '',
                'description_snippet': request.job_description or '',
                'source': 'manual',
            }

        if not job:
            raise HTTPException(status_code=404, detail="Job not found. Provide job_url, job_data, or job_description.")

        # Use provided keywords or extract from description
        description_text = job.get('description_snippet', '') or request.job_description or ''
        keywords = request.inject_keywords or extractor.extract(description_text)

        pdf_path = builder.generate(job=job, keywords=keywords, archetype=request.archetype, page_format=request.format)

        return {"success": True, "pdf_path": pdf_path, "keywords": keywords}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/resume/download/{filename}")
async def download_resume(filename: str):
    """Download a generated resume PDF"""
    # Path traversal protection: strip directory components
    safe_filename = os.path.basename(filename)
    if safe_filename != filename or '..' in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    pdf_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "resumes")
    file_path = os.path.join(pdf_dir, safe_filename)

    # Verify resolved path is within pdf_dir
    if not os.path.realpath(file_path).startswith(os.path.realpath(pdf_dir)):
        raise HTTPException(status_code=400, detail="Invalid filename")

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Resume not found")

    # Detect media type based on file extension
    if safe_filename.endswith('.html'):
        media_type = "text/html"
    else:
        media_type = "application/pdf"

    return FileResponse(file_path, media_type=media_type, filename=safe_filename)


@app.post("/api/resume/tailor")
async def tailor_resume(request: dict):
    """Tailor user's resume to match a specific job description"""
    from openai import OpenAI
    import json as _json

    job_id = request.get("job_id")
    if not job_id:
        raise HTTPException(status_code=400, detail="job_id required")

    # Get job from DB
    from core.database import get_db_session
    from models.job import Job
    db_session = next(get_db_session())
    job = db_session.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Get user profile
    if not user_profile:
        raise HTTPException(status_code=400, detail="Upload your resume first")

    # Setup AI client
    groq_key = os.getenv('GROQ_API_KEY')
    deepseek_key = os.getenv('DEEPSEEK_API_KEY')
    openai_key = os.getenv('OPENAI_API_KEY')

    if groq_key:
        client = OpenAI(api_key=groq_key, base_url="https://api.groq.com/openai/v1")
        model = "llama-3.3-70b-versatile"
    elif deepseek_key:
        client = OpenAI(api_key=deepseek_key, base_url="https://api.deepseek.com")
        model = "deepseek-chat"
    elif openai_key:
        client = OpenAI(api_key=openai_key)
        model = "gpt-4o-mini"
    else:
        raise HTTPException(status_code=500, detail="No AI API key configured")

    jd = job.description_full or job.description_snippet or ""
    profile_json = _json.dumps(user_profile, default=str)[:3000]

    prompt = f"""You are an expert resume tailor. Given the user's resume data and a job description, rewrite the resume content to better match the JD.

RULES:
- Keep EXACT same sections, structure, and approximate word count
- Do NOT invent experience or skills the user doesn't have
- Reword bullet points to use JD keywords where the user has relevant experience
- Prioritize matching skills and achievements
- Keep it truthful - only rephrase, don't fabricate

USER RESUME:
{profile_json}

JOB TITLE: {job.title}
COMPANY: {job.company}
JOB DESCRIPTION:
{jd[:2000]}

Return JSON with:
- "summary": tailored summary (2-3 lines)
- "experience": array of {{"company","title","bullets":["..."]}}
- "skills_highlighted": array of skills from user that match JD
- "keywords_added": array of JD keywords incorporated
- "keywords_missing": array of JD keywords user lacks
- "ats_score": number 0-100 estimated ATS match percentage
"""

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=3000,
        )
        raw = response.choices[0].message.content or ""
        raw = raw.strip()
        if raw.startswith("```"): raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
        if raw.endswith("```"): raw = raw[:-3]
        if raw.startswith("json"): raw = raw[4:]
        result = _json.loads(raw.strip())
        result["job_title"] = job.title
        result["company"] = job.company
        result["original_profile"] = user_profile
        return {"success": True, **result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI tailoring failed: {str(e)}")


@app.post("/api/resume/cover-letter")
async def generate_cover_letter(request: dict):
    """Generate a cover letter for a specific job"""
    from openai import OpenAI
    import json as _json

    job_id = request.get("job_id")
    if not job_id:
        raise HTTPException(status_code=400, detail="job_id required")

    from core.database import get_db_session
    from models.job import Job
    db_session = next(get_db_session())
    job = db_session.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    if not user_profile:
        raise HTTPException(status_code=400, detail="Upload your resume first")

    groq_key = os.getenv('GROQ_API_KEY')
    deepseek_key = os.getenv('DEEPSEEK_API_KEY')
    openai_key = os.getenv('OPENAI_API_KEY')

    if groq_key:
        client = OpenAI(api_key=groq_key, base_url="https://api.groq.com/openai/v1")
        model = "llama-3.3-70b-versatile"
    elif deepseek_key:
        client = OpenAI(api_key=deepseek_key, base_url="https://api.deepseek.com")
        model = "deepseek-chat"
    elif openai_key:
        client = OpenAI(api_key=openai_key)
        model = "gpt-4o-mini"
    else:
        raise HTTPException(status_code=500, detail="No AI API key configured")

    name = user_profile.get("name", "Candidate")
    skills = user_profile.get("skills", [])
    if isinstance(skills, dict):
        skills = [s for v in skills.values() if isinstance(v, list) for s in v]

    prompt = f"""Write a concise, professional cover letter (200-250 words) for:

Candidate: {name}
Key Skills: {', '.join(skills[:15]) if skills else 'Software Engineering'}
Applying for: {job.title} at {job.company}
Job Description: {(job.description_full or job.description_snippet or '')[:1500]}

Rules:
- Professional but not generic
- Reference specific skills that match the JD
- Show enthusiasm for the company/role
- Keep it concise (3-4 paragraphs)
- Don't use clichés like "I am writing to express my interest"
"""

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.5,
            max_tokens=1000,
        )
        letter = response.choices[0].message.content or ""
        return {"success": True, "cover_letter": letter.strip(), "job_title": job.title, "company": job.company}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cover letter generation failed: {str(e)}")


@app.get("/api/resume/archetypes")
async def get_archetypes():
    """List available resume archetypes"""
    return {
        "archetypes": [
            {"id": "general", "name": "General / Balanced", "description": "Well-rounded profile"},
            {"id": "full_stack", "name": "Full Stack Developer", "description": "Frontend + Backend emphasis"},
            {"id": "backend", "name": "Backend Engineer", "description": "APIs, databases, infrastructure"},
            {"id": "frontend", "name": "Frontend Developer", "description": "UI/UX, React, styling"},
            {"id": "data", "name": "Data Engineer / Scientist", "description": "Data pipelines, ML, analytics"},
            {"id": "devops", "name": "DevOps / SRE", "description": "CI/CD, cloud, infrastructure"},
        ]
    }


# ==================== Batch Processing ====================

batch_status = {"running": False, "progress": {}, "results": {}}

class BatchProcessRequest(BaseModel):
    action: str  # 'scan_and_evaluate', 'evaluate_all', 'generate_resumes'
    max_workers: Optional[int] = 8
    options: Optional[dict] = None

@app.post("/api/batch/start")
async def start_batch(request: BatchProcessRequest, background_tasks: BackgroundTasks):
    """Start batch processing"""
    global batch_status
    try:
        if batch_status["running"]:
            return {"success": False, "message": "Batch already in progress"}

        def run_batch():
            global batch_status
            batch_status = {"running": True, "progress": {"status": "starting", "completed": 0, "total": 0}, "results": {}}
            try:
                from modules.batch_conductor import BatchConductor
                conductor = BatchConductor(max_workers=request.max_workers)

                if request.action == 'scan_and_evaluate':
                    from modules.portal_scanner import PortalScanner
                    from modules.deduplication import JobDeduplicator
                    from modules.job_evaluator import JobEvaluator

                    scanner = PortalScanner()
                    deduplicator = JobDeduplicator()
                    evaluator = JobEvaluator()

                    jobs = scanner.scan_all_portals()
                    new_jobs, skipped = deduplicator.filter_new_jobs(jobs)

                    batch_status["progress"]["total"] = len(new_jobs)
                    evaluated = conductor.process_queue(
                        items=new_jobs,
                        processor_fn=evaluator.evaluate_job,
                        progress_callback=lambda p: batch_status["progress"].update(p)
                    )

                    batch_status = {
                        "running": False,
                        "progress": {"status": "completed"},
                        "results": {"total_scraped": len(jobs), "new": len(new_jobs), "skipped": skipped, "evaluated": len(evaluated)}
                    }

                elif request.action == 'evaluate_all':
                    from modules.job_evaluator import JobEvaluator
                    evaluator = JobEvaluator()
                    batch_status["progress"]["total"] = len(job_cache)
                    evaluated = conductor.process_queue(
                        items=job_cache,
                        processor_fn=evaluator.evaluate_job,
                        progress_callback=lambda p: batch_status["progress"].update(p)
                    )
                    batch_status = {
                        "running": False,
                        "progress": {"status": "completed"},
                        "results": {"evaluated": len(evaluated)}
                    }

            except Exception as e:
                batch_status = {"running": False, "progress": {"status": "error", "error": str(e)}, "results": {}}

        background_tasks.add_task(run_batch)
        return {"success": True, "message": f"Batch '{request.action}' started"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/batch/status")
async def get_batch_status():
    """Get current batch processing progress"""
    return batch_status


# ==================== Career Ops Full Pipeline ====================

class CareerOpsRequest(BaseModel):
    include_linkedin: Optional[bool] = True
    linkedin_queries: Optional[List[str]] = None
    max_results_per_query: Optional[int] = 50
    location: Optional[str] = "United States"
    auto_evaluate: Optional[bool] = True
    auto_generate_resumes: Optional[bool] = False

career_ops_status = {"running": False, "progress": {}, "results": {}}

@app.post("/api/career-ops/full-scan")
async def full_career_ops_scan(request: CareerOpsRequest, background_tasks: BackgroundTasks):
    """Run the complete career-ops pipeline: scan -> dedup -> evaluate -> generate resumes"""
    global career_ops_status, job_cache
    try:
        if career_ops_status["running"]:
            return {"success": False, "message": "Career-ops scan already in progress"}

        def run_career_ops():
            global career_ops_status, job_cache
            career_ops_status = {"running": True, "progress": {"step": "initializing"}, "results": {}}
            try:
                from modules.portal_scanner import PortalScanner
                from modules.deduplication import JobDeduplicator
                from modules.job_evaluator import JobEvaluator

                scanner = PortalScanner()
                deduplicator = JobDeduplicator()
                all_jobs = []

                # Step 1: Portal scan
                career_ops_status["progress"] = {"step": "scanning_portals", "detail": "Scanning company career pages..."}
                portal_jobs = scanner.scan_all_portals()
                all_jobs.extend(portal_jobs)

                # Step 2: LinkedIn search
                if request.include_linkedin and request.linkedin_queries:
                    career_ops_status["progress"] = {"step": "linkedin_search", "detail": f"Searching {len(request.linkedin_queries)} queries on LinkedIn..."}
                    linkedin_jobs = workflow.job_scraper.search_jobs_multi(
                        queries=request.linkedin_queries,
                        location=request.location or "United States",
                        max_results_per_query=request.max_results_per_query
                    )
                    all_jobs.extend(linkedin_jobs)

                # Step 3: Deduplicate
                career_ops_status["progress"] = {"step": "deduplication", "detail": f"Deduplicating {len(all_jobs)} jobs..."}
                new_jobs, skipped = deduplicator.filter_new_jobs(all_jobs)

                # Save to database
                for job in new_jobs:
                    try:
                        db.add_application(
                            job_title=job.get('title', 'Unknown'),
                            company=job.get('company', 'Unknown'),
                            location=job.get('location', 'Unknown'),
                            job_url=job.get('job_url', ''),
                            status='found',
                            salary_range=job.get('salary'),
                            notes=f"Source: {job.get('source', 'unknown')}"
                        )
                    except:
                        continue

                job_cache = new_jobs

                # Step 4: Evaluate
                evaluated_count = 0
                gate_passed = 0
                if request.auto_evaluate and new_jobs:
                    career_ops_status["progress"] = {"step": "evaluating", "detail": f"Evaluating {len(new_jobs)} jobs..."}
                    evaluator = JobEvaluator()
                    evaluated = evaluator.evaluate_batch(new_jobs)
                    evaluated_count = len(evaluated)
                    gate_passed = sum(1 for j in evaluated if j.get('evaluation', {}).get('gate_pass', False))

                # Step 5: Generate resumes for top jobs
                resumes_generated = 0
                if request.auto_generate_resumes and gate_passed > 0:
                    career_ops_status["progress"] = {"step": "generating_resumes", "detail": f"Generating resumes for top {min(gate_passed, 10)} jobs..."}
                    from modules.resume_generator.resume_builder import ResumeBuilder
                    from modules.resume_generator.keyword_extractor import KeywordExtractor
                    builder = ResumeBuilder()
                    extractor = KeywordExtractor()
                    top_jobs = [j for j in evaluated if j.get('evaluation', {}).get('gate_pass', False)][:10]
                    for job in top_jobs:
                        try:
                            keywords = extractor.extract(job.get('description_snippet', ''))
                            builder.generate(job=job, keywords=keywords)
                            resumes_generated += 1
                        except:
                            continue

                career_ops_status = {
                    "running": False,
                    "progress": {"step": "completed"},
                    "results": {
                        "total_scraped": len(all_jobs),
                        "portal_jobs": len(portal_jobs),
                        "linkedin_jobs": len(all_jobs) - len(portal_jobs),
                        "duplicates_skipped": skipped,
                        "new_jobs": len(new_jobs),
                        "evaluated": evaluated_count,
                        "gate_passed": gate_passed,
                        "resumes_generated": resumes_generated
                    }
                }
            except Exception as e:
                career_ops_status = {"running": False, "progress": {"step": "error", "error": str(e)}, "results": {}}

        background_tasks.add_task(run_career_ops)
        return {"success": True, "message": "Career-ops full scan started"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/career-ops/status")
async def get_career_ops_status():
    """Get status of running career-ops scan"""
    return career_ops_status


# ==================== Deduplication Stats ====================

@app.get("/api/dedup/stats")
async def get_dedup_stats():
    """Get deduplication statistics"""
    try:
        from modules.deduplication import JobDeduplicator
        deduplicator = JobDeduplicator()
        stats = deduplicator.get_stats()
        return {"success": True, "stats": stats}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Unified Job Feed ====================

@app.get("/api/jobs/feed")
async def get_unified_job_feed(
    source: Optional[str] = None,
    hours: Optional[int] = None,
    min_grade: Optional[str] = None,
    sort_by: Optional[str] = "newest",
    keyword: Optional[str] = None,
    location: Optional[str] = None
):
    """Unified job feed from all sources with filtering"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        query = '''
            SELECT sh.id, sh.job_url, sh.title, sh.company, sh.source,
                   sh.first_seen_at, sh.last_seen_at, sh.times_seen,
                   sh.evaluated, sh.evaluation_grade, sh.evaluation_score,
                   sh.applied
            FROM scan_history sh
            WHERE 1=1
        '''
        params = []

        if source and source != 'all':
            query += ' AND sh.source = ?'
            params.append(source)

        if hours:
            query += " AND sh.first_seen_at >= datetime('now', ?)"
            params.append(f'-{hours} hours')

        if min_grade:
            grade_scores = {'A': 85, 'B': 70, 'C': 55, 'D': 40}
            min_score = grade_scores.get(min_grade, 0)
            query += ' AND (sh.evaluation_score >= ? OR sh.evaluation_score IS NULL)'
            params.append(min_score)

        if keyword:
            query += ' AND (sh.title LIKE ? OR sh.company LIKE ?)'
            kw_pattern = f'%{keyword}%'
            params.extend([kw_pattern, kw_pattern])

        if location:
            # scan_history doesn't have location column, so skip for now
            pass

        if sort_by == 'newest':
            query += ' ORDER BY sh.first_seen_at DESC'
        elif sort_by == 'score':
            query += ' ORDER BY COALESCE(sh.evaluation_score, 0) DESC'
        elif sort_by == 'company':
            query += ' ORDER BY sh.company ASC'

        query += ' LIMIT 500'

        cursor.execute(query, params)
        rows = cursor.fetchall()

        jobs = []
        for row in rows:
            jobs.append({
                'id': row['id'],
                'title': row['title'],
                'company': row['company'],
                'job_url': row['job_url'],
                'source': row['source'],
                'first_seen_at': row['first_seen_at'],
                'last_seen_at': row['last_seen_at'],
                'times_seen': row['times_seen'],
                'evaluated': bool(row['evaluated']),
                'evaluation_grade': row['evaluation_grade'],
                'evaluation_score': row['evaluation_score'],
                'applied': bool(row['applied']),
            })

        # Get available sources for filter UI
        cursor.execute('SELECT source, COUNT(*) as count FROM scan_history GROUP BY source ORDER BY count DESC')
        sources = [{'name': r['source'], 'count': r['count']} for r in cursor.fetchall()]

        conn.close()

        return {
            "success": True,
            "jobs": jobs,
            "total": len(jobs),
            "sources": sources
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== JD Analysis ====================

class AnalyzeJDRequest(BaseModel):
    job_description: str
    user_skills: Optional[List[str]] = None

@app.post("/api/resume/analyze-jd")
async def analyze_job_description(request: AnalyzeJDRequest):
    """Analyze a job description for keyword coverage, ATS score, and improvement suggestions"""
    try:
        from modules.resume_generator.keyword_extractor import KeywordExtractor
        import re as re_mod
        extractor = KeywordExtractor()

        desc = request.job_description
        user_skills = request.user_skills or []

        keywords = extractor.extract(desc, user_skills=user_skills)
        categories = extractor.get_keyword_categories(keywords)

        user_skills_lower = {s.lower() for s in user_skills}
        matched_skills = [kw for kw in keywords if kw.lower() in user_skills_lower]
        missing_skills = [kw for kw in keywords if kw.lower() not in user_skills_lower]

        coverage_pct = round((len(matched_skills) / max(len(keywords), 1)) * 100)
        ats_score = min(100, coverage_pct + 15)

        suggestions = []
        if len(missing_skills) > 5:
            suggestions.append(f"Add {len(missing_skills)} missing keywords to your resume: {', '.join(missing_skills[:5])}...")
        elif missing_skills:
            suggestions.append(f"Consider adding: {', '.join(missing_skills)}")

        if 'languages' not in categories and 'backend' not in categories:
            suggestions.append("No programming languages detected - ensure your resume lists relevant languages")

        if coverage_pct < 40:
            suggestions.append("Low keyword match - heavily tailor your resume for this role")
        elif coverage_pct < 70:
            suggestions.append("Moderate match - add missing technical skills to improve ATS pass rate")
        else:
            suggestions.append("Strong keyword alignment - focus on quantifying achievements")

        exp_match = re_mod.findall(r'(\d+)\+?\s*(?:years?|yrs?)', desc.lower())
        experience_required = None
        if exp_match:
            experience_required = f"{exp_match[0]}+ years"
            suggestions.append(f"Role requires {exp_match[0]}+ years experience - highlight relevant tenure")

        return {
            "success": True,
            "keywords": keywords,
            "categories": categories,
            "keyword_categories": categories,
            "matched_skills": matched_skills,
            "missing_skills": missing_skills,
            "coverage_percent": coverage_pct,
            "keyword_coverage": {
                "matched": len(matched_skills),
                "total": len(keywords),
                "percentage": coverage_pct
            },
            "ats_score": ats_score,
            "suggestions": suggestions,
            "experience_required": experience_required,
            "total_keywords_found": len(keywords)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Email Outreach ====================

class OutreachRequest(BaseModel):
    job_url: Optional[str] = None
    recruiter_name: Optional[str] = None
    recruiter_email: Optional[str] = None
    company: Optional[str] = None
    job_title: Optional[str] = None
    template: Optional[str] = None
    template_id: Optional[str] = None  # frontend sends template_id
    subject: Optional[str] = None      # optional override from editor
    body: Optional[str] = None         # optional override from editor

    def get_template_key(self) -> str:
        return self.template_id or self.template or "introduction"

class BulkOutreachRequest(BaseModel):
    job_urls: Optional[List[str]] = None
    template: Optional[str] = None
    template_id: Optional[str] = None  # frontend sends template_id
    delay_seconds: Optional[int] = 60

    def get_template_key(self) -> str:
        return self.template_id or self.template or "introduction"

outreach_log = []

OUTREACH_TEMPLATES = {
    "introduction": {
        "name": "Introduction",
        "subject": "Excited About the {job_title} Role at {company}",
        "body": "Hi {recruiter_name},\n\nI came across the {job_title} position at {company} and I'm very excited about the opportunity. With my background in software engineering and experience with {top_skills}, I believe I'd be a strong fit for this role.\n\nI've attached my tailored resume for your review. I'd love the chance to discuss how my skills align with your team's needs.\n\nBest regards,\n{user_name}"
    },
    "follow_up": {
        "name": "Follow Up",
        "subject": "Following Up on {job_title} Application - {company}",
        "body": "Hi {recruiter_name},\n\nI wanted to follow up on my application for the {job_title} position at {company}. I remain very interested in this opportunity and would welcome the chance to discuss my qualifications further.\n\nPlease let me know if you need any additional information from me.\n\nBest regards,\n{user_name}"
    },
    "referral": {
        "name": "Referral Request",
        "subject": "Would you be open to a quick chat about {company}?",
        "body": "Hi {recruiter_name},\n\nI noticed you work at {company} and I'm very interested in the {job_title} role. I'd love to learn more about the team culture and the technical challenges you're solving.\n\nWould you be open to a brief 15-minute chat? I'd really appreciate any insights.\n\nBest regards,\n{user_name}"
    },
    "thank_you": {
        "name": "Thank You",
        "subject": "Thank You - {job_title} Discussion",
        "body": "Hi {recruiter_name},\n\nThank you so much for taking the time to discuss the {job_title} position at {company}. I really enjoyed learning about the role and the team.\n\nI'm very enthusiastic about this opportunity and look forward to the next steps.\n\nBest regards,\n{user_name}"
    }
}

def _get_user_profile_for_outreach():
    """Load user name and skills for email templates"""
    user_name = "Your Name"
    top_skills = "Python, React, AWS"
    config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "user_config.json")
    if os.path.exists(config_path):
        import json
        with open(config_path, 'r') as f:
            profile = json.load(f)
            user_name = profile.get('name', user_name)
            skills = profile.get('skills', [])
            top_skills = ', '.join(skills[:3]) if skills else top_skills
    return user_name, top_skills

@app.get("/api/outreach/templates")
async def get_outreach_templates():
    """Get email outreach templates as array with id/name/preview"""
    templates_list = [
        {"id": key, "name": tpl["name"], "preview": tpl["body"][:120] + "..."}
        for key, tpl in OUTREACH_TEMPLATES.items()
    ]
    return {"success": True, "templates": templates_list}

@app.post("/api/outreach/generate")
async def generate_outreach_email(request: OutreachRequest):
    """Generate a personalized outreach email"""
    try:
        template_key = request.get_template_key()
        template = OUTREACH_TEMPLATES.get(template_key, OUTREACH_TEMPLATES["introduction"])
        user_name, top_skills = _get_user_profile_for_outreach()

        fill = {
            "job_title": request.job_title or "Software Engineer",
            "company": request.company or "your company",
            "recruiter_name": request.recruiter_name or "Hiring Manager",
            "user_name": user_name,
            "top_skills": top_skills
        }

        subject = template["subject"].format(**fill)
        body = template["body"].format(**fill)

        return {
            "success": True,
            "subject": subject,
            "body": body,
            "template_used": template_key,
            "recruiter_name": request.recruiter_name,
            "company": request.company
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/outreach/send")
async def send_outreach_email(request: OutreachRequest):
    """Queue an outreach email"""
    try:
        # Use provided subject/body overrides from editor, or generate fresh
        if request.subject and request.body:
            subject = request.subject
            body = request.body
        else:
            email_data = await generate_outreach_email(request)
            subject = email_data["subject"]
            body = email_data["body"]

        template_key = request.get_template_key()

        log_entry = {
            "id": len(outreach_log) + 1,
            "recruiter_name": request.recruiter_name,
            "recruiter_email": request.recruiter_email,
            "company": request.company,
            "job_title": request.job_title,
            "subject": subject,
            "body": body,
            "template": template_key,
            "status": "queued",
            "created_at": datetime.now().isoformat(),
            "sent_at": None
        }
        outreach_log.append(log_entry)

        return {"success": True, "email": log_entry, "message": "Email queued for sending"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/outreach/log")
async def get_outreach_log(status: Optional[str] = None):
    """Get outreach email log"""
    log = outreach_log
    if status:
        log = [e for e in log if e["status"] == status]
    return {"success": True, "log": log, "total": len(log)}

@app.post("/api/outreach/bulk")
async def bulk_outreach(request: BulkOutreachRequest):
    """Generate and queue outreach emails for multiple jobs"""
    try:
        generated = []
        jobs_to_process = []

        if request.job_urls:
            for url in request.job_urls:
                job = next((j for j in job_cache if j.get('job_url') == url), None)
                if job:
                    jobs_to_process.append(job)
        else:
            jobs_to_process = [j for j in job_cache if j.get('recruiter_info', {}).get('name')]

        for job in jobs_to_process:
            recruiter = job.get('recruiter_info', {})
            if not recruiter.get('name'):
                continue

            template_key = request.get_template_key()
            req = OutreachRequest(
                job_url=job.get('job_url'),
                recruiter_name=recruiter.get('name', 'Hiring Manager'),
                company=job.get('company'),
                job_title=job.get('title'),
                template=template_key
            )
            result = await generate_outreach_email(req)

            log_entry = {
                "id": len(outreach_log) + 1,
                "recruiter_name": recruiter.get('name'),
                "recruiter_email": recruiter.get('email'),
                "company": job.get('company'),
                "job_title": job.get('title'),
                "job_url": job.get('job_url'),
                "subject": result["subject"],
                "body": result["body"],
                "template": template_key,
                "status": "queued",
                "created_at": datetime.now().isoformat(),
                "sent_at": None
            }
            outreach_log.append(log_entry)
            generated.append(log_entry)

        return {"success": True, "generated": len(generated), "emails": generated, "message": f"Queued {len(generated)} outreach emails"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/outreach/import")
async def import_outreach_contacts(file: UploadFile = File(...), template: Optional[str] = "introduction"):
    """Import contacts from Excel/CSV file and queue outreach emails for all"""
    try:
        import io
        contents = await file.read()

        contacts = []
        if file.filename.endswith('.csv'):
            import csv
            reader = csv.DictReader(io.StringIO(contents.decode('utf-8')))
            for row in reader:
                contacts.append({
                    'name': row.get('name', '') or row.get('Name', '') or row.get('recruiter_name', ''),
                    'email': row.get('email', '') or row.get('Email', '') or row.get('recruiter_email', ''),
                    'company': row.get('company', '') or row.get('Company', ''),
                    'job_title': row.get('job_title', '') or row.get('Job Title', '') or row.get('title', ''),
                })
        elif file.filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [str(cell.value or '').lower().strip() for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                contacts.append({
                    'name': str(row_dict.get('name', '') or row_dict.get('recruiter_name', '') or ''),
                    'email': str(row_dict.get('email', '') or row_dict.get('recruiter_email', '') or ''),
                    'company': str(row_dict.get('company', '') or ''),
                    'job_title': str(row_dict.get('job_title', '') or row_dict.get('title', '') or ''),
                })
        else:
            raise HTTPException(status_code=400, detail="Upload .csv or .xlsx file")

        # Generate and queue emails for each contact
        template_key = template or "introduction"
        tpl = OUTREACH_TEMPLATES.get(template_key, OUTREACH_TEMPLATES["introduction"])
        user_name, top_skills = _get_user_profile_for_outreach()
        queued = []

        for contact in contacts:
            if not contact.get('email'):
                continue
            fill = {
                "job_title": contact.get('job_title') or "Software Engineer",
                "company": contact.get('company') or "your company",
                "recruiter_name": contact.get('name') or "Hiring Manager",
                "user_name": user_name,
                "top_skills": top_skills,
            }
            subject = tpl["subject"].format(**fill)
            body = tpl["body"].format(**fill)

            entry = {
                "id": len(outreach_log) + len(queued) + 1,
                "recruiter_name": contact.get('name'),
                "recruiter_email": contact.get('email'),
                "company": contact.get('company'),
                "job_title": contact.get('job_title'),
                "subject": subject,
                "body": body,
                "template": template_key,
                "status": "queued",
                "created_at": datetime.now().isoformat(),
                "sent_at": None,
            }
            queued.append(entry)

        outreach_log.extend(queued)

        return {
            "success": True,
            "imported": len(contacts),
            "queued": len(queued),
            "skipped": len(contacts) - len(queued),
            "message": f"Imported {len(contacts)} contacts, queued {len(queued)} emails"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ATS Detection & Import ====================

ats_detection_status = {"running": False, "progress": {}, "results": {}}

class ATSDetectRequest(BaseModel):
    max_workers: Optional[int] = 10

@app.post("/api/portals/detect-ats")
async def detect_ats_for_career_pages(request: ATSDetectRequest, background_tasks: BackgroundTasks):
    """Run ATS detection on all awesome-career-pages companies"""
    global ats_detection_status
    try:
        if ats_detection_status["running"]:
            return {"success": False, "message": "Detection already in progress"}

        def run_detection():
            global ats_detection_status
            ats_detection_status = {"running": True, "progress": {"status": "starting"}, "results": {}}
            try:
                from modules.ats_detector import ATSDetector
                detector = ATSDetector()
                companies = detector.load_awesome_career_pages()

                if not companies:
                    ats_detection_status = {"running": False, "progress": {"status": "error", "error": "No companies found in Portal.json"}, "results": {}}
                    return

                def progress_cb(p):
                    ats_detection_status["progress"] = {
                        "status": "detecting",
                        "completed": p.get('completed', 0),
                        "total": p.get('total', 0),
                        "current": p.get('current', ''),
                        "current_ats": p.get('ats', ''),
                    }

                summary = detector.detect_and_save(companies, max_workers=request.max_workers, progress_callback=progress_cb)
                ats_detection_status = {"running": False, "progress": {"status": "completed"}, "results": summary}
            except Exception as e:
                ats_detection_status = {"running": False, "progress": {"status": "error", "error": str(e)}, "results": {}}

        background_tasks.add_task(run_detection)
        return {"success": True, "message": "ATS detection started"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/portals/detect-ats/status")
async def get_ats_detection_status():
    """Get ATS detection progress"""
    return ats_detection_status

@app.get("/api/portals/detected")
async def get_detected_portals(ats: Optional[str] = None):
    """Get list of auto-detected portals"""
    try:
        from modules.ats_detector import ATSDetector
        detector = ATSDetector()
        portals = detector.get_detected_portals(ats_filter=ats)

        # Summary
        ats_counts = {}
        for p in portals:
            a = p.get('ats', 'unknown')
            ats_counts[a] = ats_counts.get(a, 0) + 1

        return {
            "success": True,
            "portals": portals[:200],  # Limit response size
            "total": len(portals),
            "ats_breakdown": ats_counts,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class ExtendedScanRequest(BaseModel):
    keyword_filter: Optional[str] = None
    location_filter: Optional[str] = None
    include_detected: Optional[bool] = True
    detected_ats_types: Optional[List[str]] = None

@app.post("/api/portals/scan-extended")
async def start_extended_portal_scan(request: ExtendedScanRequest, background_tasks: BackgroundTasks):
    """Scan all portals INCLUDING auto-detected ones from awesome-career-pages"""
    global portal_scan_status
    try:
        if portal_scan_status["running"]:
            return {"success": False, "message": "Scan already in progress"}

        def run_scan():
            global portal_scan_status
            portal_scan_status = {"running": True, "progress": {"status": "starting"}, "results": []}
            try:
                from modules.portal_scanner import PortalScanner
                from modules.deduplication import JobDeduplicator
                scanner = PortalScanner()
                deduplicator = JobDeduplicator()

                jobs = scanner.scan_all_portals_extended(
                    progress_callback=lambda p: portal_scan_status.update({"progress": p}),
                    keyword_filter=request.keyword_filter,
                    location_filter=request.location_filter,
                    include_detected=request.include_detected,
                    detected_ats_types=request.detected_ats_types,
                )
                new_jobs, skipped = deduplicator.filter_new_jobs(jobs)

                portal_scan_status = {
                    "running": False,
                    "progress": {"status": "completed"},
                    "results": {
                        "total_scraped": len(jobs),
                        "new_jobs": len(new_jobs),
                        "duplicates_skipped": skipped,
                    }
                }
            except Exception as e:
                portal_scan_status = {"running": False, "progress": {"status": "error", "error": str(e)}, "results": []}

        background_tasks.add_task(run_scan)
        return {"success": True, "message": "Extended portal scan started (includes auto-detected portals)"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Reddit Scraper ====================

class RedditScanRequest(BaseModel):
    subreddits: Optional[List[str]] = None
    keyword_filter: Optional[str] = None

@app.post("/api/reddit/scan")
async def scan_reddit(request: RedditScanRequest, background_tasks: BackgroundTasks):
    """Scan Reddit job subreddits"""
    try:
        from modules.board_scrapers.reddit_scraper import RedditScraper
        from modules.deduplication import JobDeduplicator

        scraper = RedditScraper()
        deduplicator = JobDeduplicator()

        subs = request.subreddits or [
            'hiring', 'forhire', 'remotejobs', 'remotework', 'WorkOnline',
            'webdev', 'freelance', 'datascience', 'digitalnomad', 'jobbit'
        ]

        jobs = scraper.scrape_all_subreddits(
            subreddits=subs,
            keyword_filter=request.keyword_filter
        )

        new_jobs, skipped = deduplicator.filter_new_jobs(jobs)

        return {
            "success": True,
            "total_scraped": len(jobs),
            "new_jobs": len(new_jobs),
            "duplicates_skipped": skipped,
            "jobs": [{"title": j.get("title"), "company": j.get("company"), "url": j.get("job_url"), "source": j.get("source")} for j in new_jobs[:20]]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Dashboard Source Stats ====================

@app.get("/api/dashboard/source-stats")
async def get_source_stats():
    """Get job stats broken down by source for dashboard"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT source, COUNT(*) as count,
                   SUM(CASE WHEN evaluated = 1 THEN 1 ELSE 0 END) as evaluated,
                   SUM(CASE WHEN applied = 1 THEN 1 ELSE 0 END) as applied,
                   SUM(CASE WHEN date(first_seen_at) = date('now') THEN 1 ELSE 0 END) as today
            FROM scan_history
            GROUP BY source ORDER BY count DESC
        ''')
        by_source = []
        for row in cursor.fetchall():
            by_source.append({
                'source': row['source'],
                'total': row['count'],
                'evaluated': row['evaluated'],
                'applied': row['applied'],
                'new_today': row['today']
            })

        cursor.execute("SELECT COUNT(*) as c FROM scan_history WHERE first_seen_at >= datetime('now', '-1 hours')")
        last_1h = cursor.fetchone()['c']
        cursor.execute("SELECT COUNT(*) as c FROM scan_history WHERE first_seen_at >= datetime('now', '-6 hours')")
        last_6h = cursor.fetchone()['c']
        cursor.execute("SELECT COUNT(*) as c FROM scan_history WHERE first_seen_at >= datetime('now', '-24 hours')")
        last_24h = cursor.fetchone()['c']
        cursor.execute("SELECT COUNT(*) as c FROM scan_history")
        total = cursor.fetchone()['c']

        conn.close()

        return {
            "success": True,
            "by_source": by_source,
            "time_stats": {
                "last_1h": last_1h,
                "last_6h": last_6h,
                "last_24h": last_24h,
                "total": total
            },
            "outreach_queued": len([e for e in outreach_log if e["status"] == "queued"]),
            "outreach_sent": len([e for e in outreach_log if e["status"] == "sent"])
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Job Tracking (Applied/Ignored/Clicked) ====================

class UpdateJobStatusRequest(BaseModel):
    job_url: str
    status: str  # 'applied', 'ignored', 'clicked', 'saved', 'interview', 'rejected', 'offer'
    notes: Optional[str] = None

@app.post("/api/jobs/track")
async def update_job_status(request: UpdateJobStatusRequest):
    """Update job tracking status (applied, ignored, clicked, etc.)"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        now = datetime.now().isoformat()

        # Update scan_history
        if request.status == 'applied':
            cursor.execute('UPDATE scan_history SET applied = 1 WHERE job_url = ?', (request.job_url,))

        # Also update/insert into a tracking table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS job_tracking (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_url TEXT NOT NULL UNIQUE,
                status TEXT NOT NULL,
                notes TEXT,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute('''
            INSERT INTO job_tracking (job_url, status, notes, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(job_url) DO UPDATE SET status = ?, notes = ?, updated_at = ?
        ''', (request.job_url, request.status, request.notes, now, request.status, request.notes, now))

        conn.commit()
        conn.close()
        return {"success": True, "status": request.status}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/jobs/tracking")
async def get_job_tracking(status: Optional[str] = None):
    """Get all tracked jobs with their statuses"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS job_tracking (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_url TEXT NOT NULL UNIQUE,
                status TEXT NOT NULL,
                notes TEXT,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        if status:
            cursor.execute('''
                SELECT jt.*, sh.title, sh.company, sh.source
                FROM job_tracking jt
                LEFT JOIN scan_history sh ON jt.job_url = sh.job_url
                WHERE jt.status = ?
                ORDER BY jt.updated_at DESC
            ''', (status,))
        else:
            cursor.execute('''
                SELECT jt.*, sh.title, sh.company, sh.source
                FROM job_tracking jt
                LEFT JOIN scan_history sh ON jt.job_url = sh.job_url
                ORDER BY jt.updated_at DESC
            ''')

        rows = cursor.fetchall()
        conn.close()

        tracked = []
        for row in rows:
            tracked.append({
                'job_url': row['job_url'],
                'status': row['status'],
                'notes': row['notes'],
                'updated_at': row['updated_at'],
                'title': row['title'] if 'title' in row.keys() else None,
                'company': row['company'] if 'company' in row.keys() else None,
                'source': row['source'] if 'source' in row.keys() else None,
            })

        return {"success": True, "tracked": tracked, "total": len(tracked)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/jobs/export-feed")
async def export_feed_to_excel(
    source: Optional[str] = None,
    status: Optional[str] = None
):
    """Export job feed to Excel with tracking status"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        query = '''
            SELECT sh.title, sh.company, sh.source, sh.job_url,
                   sh.first_seen_at, sh.evaluation_grade, sh.evaluation_score,
                   COALESCE(jt.status, 'new') as tracking_status,
                   jt.notes as tracking_notes
            FROM scan_history sh
            LEFT JOIN job_tracking jt ON sh.job_url = jt.job_url
            WHERE 1=1
        '''
        params = []
        if source and source != 'all':
            query += ' AND sh.source = ?'
            params.append(source)
        if status:
            query += ' AND COALESCE(jt.status, ?) = ?'
            params.extend(['new', status])

        query += ' ORDER BY sh.first_seen_at DESC LIMIT 1000'
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()

        # Generate Excel
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Feed Export"

        headers = ['Title', 'Company', 'Source', 'URL', 'First Seen', 'Grade', 'Score', 'Status', 'Notes']
        ws.append(headers)

        for row in rows:
            ws.append([
                row['title'], row['company'], row['source'], row['job_url'],
                row['first_seen_at'], row['evaluation_grade'],
                row['evaluation_score'], row['tracking_status'],
                row['tracking_notes']
            ])

        # Save to temp file
        export_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "exports")
        os.makedirs(export_dir, exist_ok=True)
        filename = f"job_feed_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(export_dir, filename)
        wb.save(filepath)

        return FileResponse(filepath, filename=filename,
                          media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Auto-Apply Pipeline ====================

class AutoApplyRequest(BaseModel):
    job_url: str
    auto_generate_resume: Optional[bool] = True
    archetype: Optional[str] = "auto"

@app.post("/api/jobs/auto-apply-pipeline")
async def auto_apply_pipeline(request: AutoApplyRequest):
    """Full pipeline: Read JD -> Analyze -> Generate tailored resume -> Return for apply"""
    try:
        from modules.resume_generator.resume_builder import ResumeBuilder
        from modules.resume_generator.keyword_extractor import KeywordExtractor

        builder = ResumeBuilder()
        extractor = KeywordExtractor()

        # Step 1: Find the job (from cache or scan_history)
        job = next((j for j in job_cache if j.get('job_url') == request.job_url), None)

        if not job:
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM scan_history WHERE job_url = ?', (request.job_url,))
            row = cursor.fetchone()
            conn.close()
            if row:
                job = dict(row)

        if not job:
            raise HTTPException(status_code=404, detail="Job not found in cache or database")

        # Step 2: Extract keywords from JD
        description = job.get('description_snippet', '') or job.get('title', '')
        keywords = extractor.extract(description)
        categories = extractor.get_keyword_categories(keywords)

        # Step 3: Load user profile and calculate match
        user_profile = builder.profile
        user_skills = user_profile.get('skills', [])
        user_skills_lower = {s.lower() for s in user_skills}
        matched = [kw for kw in keywords if kw.lower() in user_skills_lower]
        missing = [kw for kw in keywords if kw.lower() not in user_skills_lower]
        coverage = round((len(matched) / max(len(keywords), 1)) * 100)

        # Step 4: Generate tailored resume
        pdf_path = None
        if request.auto_generate_resume:
            pdf_path = builder.generate(
                job=job,
                keywords=keywords,
                archetype=request.archetype or 'auto',
                page_format='Letter'
            )

        # Step 5: Extract filename for download
        resume_filename = None
        if pdf_path:
            resume_filename = os.path.basename(pdf_path)

        return {
            "success": True,
            "job": {
                "title": job.get('title'),
                "company": job.get('company'),
                "location": job.get('location', ''),
                "url": job.get('job_url'),
                "source": job.get('source', 'unknown'),
                "description": description[:500]
            },
            "analysis": {
                "keywords_found": len(keywords),
                "keywords": keywords,
                "categories": categories,
                "matched_skills": matched,
                "missing_skills": missing,
                "coverage_percent": coverage,
                "ats_score": min(100, coverage + 15)
            },
            "resume": {
                "generated": bool(pdf_path),
                "filename": resume_filename,
                "download_url": f"/api/resume/download/{resume_filename}" if resume_filename else None
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== LinkedIn Hacks & Tricks ====================

class GroupMessageRequest(BaseModel):
    group_url: Optional[str] = None
    target_name: Optional[str] = None
    message: Optional[str] = None

class SavedSearchRequest(BaseModel):
    name: str
    url: str
    keywords: Optional[str] = None
    notes: Optional[str] = None

class GoogleDorkRequest(BaseModel):
    company: Optional[str] = None
    job_title: Optional[str] = None
    location: Optional[str] = None

class RecruiterEngageRequest(BaseModel):
    recruiter_profile_url: str
    strategy: Optional[str] = "comment"  # comment, tag, share
    job_url: Optional[str] = None
    message: Optional[str] = None

# In-memory saved searches store
saved_searches = []

@app.get("/api/hacks/saved-searches")
async def get_saved_searches():
    """Get all saved LinkedIn search URLs (unlimited search alert hack)"""
    return {"success": True, "searches": saved_searches, "total": len(saved_searches)}

@app.post("/api/hacks/saved-searches")
async def add_saved_search(request: SavedSearchRequest):
    """Save a LinkedIn search URL for unlimited search alerts"""
    entry = {
        "id": len(saved_searches) + 1,
        "name": request.name,
        "url": request.url,
        "keywords": request.keywords,
        "notes": request.notes,
        "created_at": datetime.now().isoformat(),
        "last_checked": None,
        "results_count": None
    }
    saved_searches.append(entry)
    return {"success": True, "search": entry}

@app.delete("/api/hacks/saved-searches/{search_id}")
async def delete_saved_search(search_id: int):
    """Delete a saved search"""
    global saved_searches
    saved_searches = [s for s in saved_searches if s["id"] != search_id]
    return {"success": True}

@app.post("/api/hacks/google-dork")
async def generate_google_dork(request: GoogleDorkRequest):
    """Generate Google search operators to find hidden LinkedIn profiles"""
    queries = []

    base = 'site:linkedin.com/in'

    if request.company and request.job_title:
        queries.append(f'{base} "{request.company}" "{request.job_title}"')
        queries.append(f'{base} "{request.company}" "{request.job_title}" "{request.location or ""}"'.strip())
    elif request.company:
        queries.append(f'{base} "{request.company}"')
        queries.append(f'{base} "{request.company}" recruiter')
        queries.append(f'{base} "{request.company}" "hiring manager"')
        queries.append(f'{base} "{request.company}" "talent acquisition"')
    elif request.job_title:
        queries.append(f'{base} "{request.job_title}"')
        if request.location:
            queries.append(f'{base} "{request.job_title}" "{request.location}"')

    # Add advanced variants
    if request.company:
        queries.append(f'site:linkedin.com/company/{request.company.lower().replace(" ", "-")}/people')
        queries.append(f'{base} "{request.company}" (CTO OR VP OR Director OR "Engineering Manager")')

    # Clean up empty quotes
    queries = [q.replace('""', '').strip() for q in queries if q.strip()]

    return {
        "success": True,
        "queries": queries,
        "tip": "Open these in incognito for best results. You can often see more profile info when not logged in.",
        "company": request.company,
        "job_title": request.job_title
    }

@app.post("/api/hacks/recruiter-engage")
async def plan_recruiter_engagement(request: RecruiterEngageRequest):
    """Generate a recruiter engagement strategy (comment/tag/share approach)"""
    strategies = {
        "comment": {
            "action": "Comment on their recent posts with valuable insights",
            "steps": [
                f"Visit {request.recruiter_profile_url} and check their recent activity",
                "Find their latest post about hiring, industry trends, or company culture",
                "Write a thoughtful comment (NOT just 'Interested!' or 'Great post!')",
                "Add value: share a relevant stat, personal experience, or thoughtful question",
                "Do this 2-3 times over a week before reaching out via DM"
            ],
            "sample_comments": [
                "This resonates with my experience building [X]. The challenge of [Y] is real — I found that [specific approach] helped tremendously.",
                "Great insight on [topic]. I've seen similar patterns in my work with [relevant experience]. Would love to hear more about how your team approaches [specific challenge].",
                "Thanks for sharing this. As someone actively working in [relevant field], I'd add that [complementary insight]. The [specific tool/approach] you mentioned is game-changing."
            ]
        },
        "tag": {
            "action": "Tag them in a relevant post you create",
            "steps": [
                "Create a LinkedIn post about a relevant topic (project, insight, achievement)",
                f"Tag the recruiter naturally: 'Inspired by a post from @recruiter about [topic]...'",
                "Make the post genuinely valuable — not a thinly veiled job ask",
                "This notifies them and puts your profile on their radar"
            ],
            "sample_posts": [
                "Just shipped [project/feature] using [tech stack]. The biggest challenge was [X] — here's how we solved it. cc @recruiter who posted about similar challenges at [company].",
                "Reflecting on [industry trend]. After [X years] in [field], I've learned that [insight]. Would love to hear perspectives from folks like @recruiter at [company]."
            ]
        },
        "share": {
            "action": "Share their job post with a personal note",
            "steps": [
                f"Find their latest job posting",
                "Share it on your feed with a genuine note about why the role excites you",
                "This directly notifies the recruiter and shows proactive interest",
                "Tag the recruiter and the company page in your share"
            ],
            "sample_share": "Excited about this [job_title] role at [company]! The focus on [specific aspect from JD] aligns perfectly with my work in [relevant area]. @recruiter — would love to chat about this opportunity."
        }
    }

    strategy = strategies.get(request.strategy, strategies["comment"])

    return {
        "success": True,
        "strategy": request.strategy,
        "recruiter_url": request.recruiter_profile_url,
        "plan": strategy,
        "general_tips": [
            "Engage 2-3 times before sending a cold DM",
            "Always provide value — never just 'Interested!'",
            "Personalize every interaction based on their content",
            "Connect the dots between your skills and their needs",
            "Be consistent but not spammy — 2-3 interactions per week max"
        ]
    }

@app.post("/api/hacks/group-message")
async def plan_group_message(request: GroupMessageRequest):
    """Plan a group-based messaging strategy to reach non-connections"""
    recommended_groups = [
        "Software Engineers", "Tech Careers", "Remote Work",
        "Python Developers", "JavaScript Developers", "React Developers",
        "DevOps & Cloud", "Data Science", "Product Management",
        "Startup Founders", "Hiring & Recruiting", "Tech Leadership"
    ]

    return {
        "success": True,
        "strategy": {
            "name": "Group Message Hack",
            "description": "Message anyone who shares a LinkedIn Group with you — no InMail needed",
            "steps": [
                "Search for industry-specific LinkedIn Groups (e.g., 'Software Engineers', 'React Developers')",
                "Join 5-10 relevant groups where recruiters and hiring managers are active",
                "Wait for approval (some groups approve instantly)",
                "Open the group → Members tab → Find your target person",
                "Click their name → 'Message' button is now available (free!)",
                "Send a personalized message referencing the shared group"
            ],
            "recommended_groups": recommended_groups,
            "sample_message": request.message or f"Hi {request.target_name or '[Name]'},\n\nI noticed we're both in [Group Name] — I've been following the discussions on [topic]. I saw that [Company] is hiring for [Role] and I'd love to chat about how my experience with [skills] might be a fit.\n\nWould you be open to a quick 10-minute call?\n\nBest regards"
        },
        "tip": "This bypasses InMail restrictions completely. Recruiters in hiring-focused groups are usually very responsive."
    }

@app.post("/api/hacks/refresh-job-post")
async def refresh_job_post_hack(job_url: str = ""):
    """Explain the 86400 job post freshness hack"""
    return {
        "success": True,
        "hack": {
            "name": "Job Post Freshness Hack (86400)",
            "description": "LinkedIn job posts expire after 30 days. The 86400 parameter (seconds in a day) can be used to manipulate post freshness.",
            "how_it_works": [
                "86400 = number of seconds in 1 day",
                "LinkedIn uses time-based parameters in job post URLs",
                "Editing or re-sharing a job post URL resets its visibility",
                "Filtering by 'Past 24 hours' uses this parameter under the hood"
            ],
            "for_job_seekers": [
                "Always filter jobs by 'Past 24 hours' or 'Past week' for freshest listings",
                "Jobs posted < 24h ago have 3x higher response rates",
                "Sort by 'Most recent' to catch just-posted roles",
                "Set up our Saved Searches to check daily"
            ],
            "original_url": job_url,
            "tip": "Apply within the first 24 hours of a job posting for the highest chance of being seen."
        }
    }

