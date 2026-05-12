"""Excel Export Service for Job Data"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from pathlib import Path
from typing import List, Dict


class ExcelExportService:
    def __init__(self):
        self.export_dir = Path(__file__).parent.parent.parent / "exports"
        self.export_dir.mkdir(exist_ok=True)

    def export_jobs_to_excel(self, jobs: List[Dict], filename: str = None) -> str:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"linkedin_jobs_{timestamp}.xlsx"

        filepath = self.export_dir / filename
        wb = Workbook()
        ws_jobs = wb.active
        ws_jobs.title = "Job Listings"
        ws_recruiters = wb.create_sheet("Recruiters & Contacts")

        # Job headers
        headers = ["No.", "Job Title", "Company", "Location", "Salary", "Posted Date", "Applicants", "Easy Apply", "Job URL"]
        for c, h in enumerate(headers, 1):
            cell = ws_jobs.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

        # Job rows
        for idx, job in enumerate(jobs, 1):
            row = [
                idx,
                job.get("title", "N/A"),
                job.get("company", "N/A"),
                job.get("location", "N/A"),
                job.get("salary", "Not specified"),
                job.get("posted_date", "N/A"),
                job.get("applicant_count", "N/A"),
                "Yes" if job.get("easy_apply") else "No",
                job.get("job_url", ""),
            ]
            for c, val in enumerate(row, 1):
                cell = ws_jobs.cell(row=idx + 1, column=c, value=val)
                if isinstance(val, str) and val.startswith("http"):
                    cell.hyperlink = val
                    cell.font = Font(color="0563C1", underline="single")
                    cell.value = "Click"

        # Recruiter headers
        rec_headers = ["Job Title", "Company", "Recruiter Name", "Recruiter Title", "Profile URL", "DM Link", "Email"]
        for c, h in enumerate(rec_headers, 1):
            cell = ws_recruiters.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

        r_row = 2
        for job in jobs:
            recruiter = job.get("recruiter_info", {})
            if recruiter and recruiter.get("name"):
                ws_recruiters.cell(row=r_row, column=1, value=job.get("title", "N/A"))
                ws_recruiters.cell(row=r_row, column=2, value=job.get("company", "N/A"))
                ws_recruiters.cell(row=r_row, column=3, value=recruiter.get("name", "N/A"))
                ws_recruiters.cell(row=r_row, column=4, value=recruiter.get("title", ""))
                profile_url = recruiter.get("profile_url", "")
                ws_recruiters.cell(row=r_row, column=5, value=profile_url)
                if profile_url and profile_url.startswith("http"):
                    ws_recruiters.cell(row=r_row, column=5).hyperlink = profile_url
                    ws_recruiters.cell(row=r_row, column=5).font = Font(color="0563C1", underline="single")
                dm_link = recruiter.get("dm_link", "")
                ws_recruiters.cell(row=r_row, column=6, value=dm_link)
                if dm_link and dm_link.startswith("http"):
                    ws_recruiters.cell(row=r_row, column=6).hyperlink = dm_link
                    ws_recruiters.cell(row=r_row, column=6).font = Font(color="0563C1", underline="single")
                ws_recruiters.cell(row=r_row, column=7, value=recruiter.get("email", ""))
                r_row += 1

            # Also add people who can help
            for person in job.get("people_who_can_help", []):
                if person.get("name"):
                    ws_recruiters.cell(row=r_row, column=1, value=job.get("title", "N/A"))
                    ws_recruiters.cell(row=r_row, column=2, value=job.get("company", "N/A"))
                    ws_recruiters.cell(row=r_row, column=3, value=person.get("name", ""))
                    ws_recruiters.cell(row=r_row, column=4, value=person.get("title", ""))
                    p_url = person.get("profile_url", "")
                    ws_recruiters.cell(row=r_row, column=5, value=p_url)
                    if p_url and p_url.startswith("http"):
                        ws_recruiters.cell(row=r_row, column=5).hyperlink = p_url
                        ws_recruiters.cell(row=r_row, column=5).font = Font(color="0563C1", underline="single")
                    p_dm = person.get("dm_link", "")
                    ws_recruiters.cell(row=r_row, column=6, value=p_dm)
                    if p_dm and p_dm.startswith("http"):
                        ws_recruiters.cell(row=r_row, column=6).hyperlink = p_dm
                        ws_recruiters.cell(row=r_row, column=6).font = Font(color="0563C1", underline="single")
                    r_row += 1

        wb.save(filepath)
        return str(filepath)
